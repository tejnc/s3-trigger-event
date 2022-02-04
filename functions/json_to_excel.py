import os
from json import loads

import boto3
from openpyxl import Workbook

from models.payment import Payment
from utils.constants import SUCCESS_CREATED
from utils.mongo import db_config
from utils.response import success_response

# temp_folder_name = "/tmp"
# s3 = boto3.resource("s3")


def json_to_excel(json_file):
    wb = Workbook()
    worksheet_01 = wb.active
    worksheet_01.title = "Payment data"

    # configuring body
    worksheet_01.cell(1, 1, "Amount")
    worksheet_01.cell(1, 2, "Medium")
    worksheet_01.cell(1, 3, "Created_at")
    worksheet_01.cell(1, 4, "Years")
    worksheet_01.cell(1, 5, "Pay_for")

    row = 1
    for i in range(len(json_file)):
        row += 1
        worksheet_01.cell(row, 1, json_file[i]["amount"])
        worksheet_01.cell(row, 2, json_file[i]["medium"])
        worksheet_01.cell(row, 3, json_file[i]["created_at"])
        worksheet_01.cell(row, 4, json_file[i]["years"])
        worksheet_01.cell(row, 5, json_file[i]["pay_for"])

    # os.makedirs(temp_folder_name, exist_ok=True)
    # wb.save("/tmp/payment_data.xlsx")

    wb.save("/home/tej/Documents/Intern/payment_data.xlsx")
    file_path = "/home/tej/Documents/Intern/payment_data.xlsx"
    # file_path = "/tmp/payment_data.xlsx"
    return file_path


def main(event, context):
    db_config()

    user: Payment = Payment.objects()
    user_json = loads(user.to_json())
    excel_path = json_to_excel(json_file=user_json)

    # s3_client = get_service_instance(event,"s3")

    # file_name = excel_path.split("/")[-1]
    # object_name = "excel/{}".format(file_name)
    # bucket = "test-demo15"

    # s3.meta.client.upload_file(excel_path, bucket, object_name) #while using s3 resource

    # s3_client.upload_file(excel_path,bucket, object_name) # while using s3 client

    return success_response(
        status_code=200, message="working", data={"path": excel_path}
    )
